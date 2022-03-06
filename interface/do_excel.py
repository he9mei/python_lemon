#!/usr/bin/python3
# -*- coding: utf-8 -*-

# 2-2读取excel
import openpyxl
from interface.http_request import HTTPRequest


# 打开文件
# cases_file = openpyxl.load_workbook("E:\\cases.xlsx")
# 不支持旧的xls格式
# 遇到问题：raise IOError("File contains no valid workbook part")
# 原因xls文件改成xlsx后缀
# 解决：重新创建xlsx文件


class DoExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_cases(self):
        workbook = openpyxl.load_workbook(self.file_name)
        sheet = workbook[self.sheet_name]
        cases = []
        max_row = sheet.max_row
        for r in range(2, max_row + 1):
            case = {}
            case["case_id"] = sheet.cell(row=r, column=1).value
            case["title"] = sheet.cell(row=r, column=2).value
            case["url"] = sheet.cell(row=r, column=3).value
            case["data"] = sheet.cell(row=r, column=4).value
            case["method"] = sheet.cell(row=r, column=5).value
            case["expected"] = sheet.cell(row=r, column=6).value
            cases.append(case)
        workbook.save(self.file_name)
        workbook.close()
        return cases

    def write_result(self, row, actual, result):
        workbook = openpyxl.load_workbook(self.file_name)
        sheet = workbook[self.sheet_name]
        sheet.cell(row, column=7).value = actual
        sheet.cell(row, column=7).value = result
        workbook.save(self.file_name)
        workbook.close()


if __name__ == "__main__":
    excel = DoExcel("E:\\cases.xlsx", "login")
    cases = excel.get_cases()
    for case in cases:
        data = eval(case["data"])  # 将字符串转换成字典格式（存放在excel中的都是当做字符串）
        request = HTTPRequest(method=case["method"], url=case["url"], data=data)
        actual = request.get_text()
        if actual == case["expected"]:
            result = "PASS"
        else:
            result = "FAIL"
        excel.write_result(case["case_id"] + 1, actual, result)

    print("执行完毕！")

# 如果文件本身是打开的会报错，需要先关闭
# 执行成功，视频看到2-2的50分钟