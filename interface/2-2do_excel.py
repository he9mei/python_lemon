#!/usr/bin/python3
# -*- coding: utf-8 -*-

# 2-2读取excel
import openpyxl
from interface.http_request import HTTPRequest


# 打开文件
# cases_file = openpyxl.load_workbook("E:\\cases.xlsx")
# 不支持旧的xls格式
# 遇到问题：raise IOError("File contains no valid workbook part")
# 原因xls文件改成xlsx后缀
# 解决：重新创建xlsx文件


class DoExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_cases(self):
        workbook = openpyxl.load_workbook(self.file_name)
        sheet = workbook[self.sheet_name]
        cases = []
        max_row = sheet.max_row
        for r in range(2, max_row + 1):
            case = {}
            case["case_id"] = sheet.cell(row=r, column=1).value
            case["title"] = sheet.cell(row=r, column=2).value
            case["url"] = sheet.cell(row=r, column=3).value
            case["data"] = sheet.cell(row=r, column=4).value
            case["method"] = sheet.cell(row=r, column=5).value
            case["expected"] = sheet.cell(row=r, column=6).value
            cases.append(case)
        workbook.save(self.file_name)
        workbook.close()
        return cases

    def write_result(self, row, actual, result):
        workbook = openpyxl.load_workbook(self.file_name)
        sheet = workbook[self.sheet_name]
        sheet.cell(row, column=7).value = actual
        sheet.cell(row, column=8).value = result
        workbook.save(self.file_name)
        workbook.close()


if __name__ == "__main__":
    excel = DoExcel("E:\\cases.xlsx", "login")
    cases = excel.get_cases()

    for case in cases:
        print("现在运行的测试用例：{}".format(case["title"]))
        data = eval(case["data"])  # 此处存放在excel的data是字符串，需要转换成字典
        request = HTTPRequest(method=case["method"], url=case["url"], data=data)
        actual = request.get_text()
        if actual == case["expected"]:
            result = "PASS"
        else:
            result = "FAIL"
        excel.write_result(case["case_id"] + 1, actual, result)  # 此处Excel中的case_id是int

    print("执行完毕！")

# 如果文件本身是打开的会报错，需要先关闭


'''
问题：push失败
-->git config --global credential.helper store
-->git push
he9mei
8uhb*UHBhhm
错误：Support for password authentication was removed on August 13, 2021. Please use a personal access token instead
需要token代替密码登录,如何生成token？
进入Git后，头像下方Settings-->Developer settings-->personal access token
填写note，选择90天，勾选repo，Generate token-->
https://github.com/settings/tokens
token：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
再次push时，密码使用token即可
参考：https://blog.csdn.net/weixin_41010198/article/details/119698015
'''

