#!/usr/bin/python3
# -*- coding: utf-8 -*-


# ---老师的讲解---自己没有这样写（5_2练习），参考性很高---

from openpyxl import load_workbook

class MyExcel:

    def __init__(self,filename):
        # 打开工作薄
        try:
            self.__wb = load_workbook(filename)
        except:
            print("要操作的excel不存在，请检查！！")
            raise # 抛出异常

    def select_sheet_by_name(self,sheet_name):
        # # 判断表单名称是否存在
        # if sheet_name in self.__wb.sheetnames:
        #     self.sh = self.__wb[sheet_name]
        # else:
        #     print("表单名称不存在，请确认表单名称正确！！")
        try:
            self.sh = self.__wb[sheet_name]
        except:
            print("表单名称不存在，请确认表单名称正确！！")
            raise

    def get_cell_data_by_row_and_column(self,row,col):
        print(type(row))   # 调试  debug   更细致。
        if type(row) is int and type(col) is int:
            if row >=1 and col >=1:
                print("读取了{}行{}列的数据。")  # info
                return self.sh.cell(row,col).value
            else:
                print("下标必须大于1")   # error
        else:
            print("参数类型错误！行号和列号应该为整数。")  # error

    # 获取 一行数据
    def get_row_datas_by_row_num(self,row):
        # 整行数据
        row_datas =[]
        if type(row) is int:
            # 遍历列
            for col in range(1,self.sh.max_column+1):
                row_datas.append(self.sh.cell(row,col).value)
        return row_datas

    def get_all_datas(self):
        # 所有数据
        all_datas = []
        # 遍历行，调用上一个函数。
        for row in range(1,self.sh.max_row +1):
            row_datas = self.get_row_datas_by_row_num(row)
            all_datas.append(row_datas)
        return all_datas

    # 更新单无格的数据
    def update_cell_data_by_row_and_column(self,row,col,new_value):
        if type(row) is int and type(col) is int:
            if row >= 1 and col >= 1:
                self.sh.cell(row, col).value = new_value

    # 保存数据
    def save_datas_to_file(self,filename):
        try:
            self.__wb.save(filename)
        except PermissionError:
            print("文件权限被占用，无法保存。保存失败！！")
            raise
        except FileNotFoundError:
            print("文件路径不存在，请检查！！")
            raise
        except:
            print("文件保存失败，请检查原因！！！")
            raise

# 你的使用目标不明确。没有明显的应用场景。

if __name__ == "__main__":
    # 测试
    me = MyExcel("datas.xlsx")  # 加载文件
    me.select_sheet_by_name("case_datas")  # 选择表单
    datas = me.get_row_datas_by_row_num(3) # 读一行数据
    print(datas)
    all = me.get_all_datas()   # 所有数据
    print(all)

    me.update_cell_data_by_row_and_column(7,7,"我偷偷进来了！！！")
    me.update_cell_data_by_row_and_column(9,9,"同学们，你们类和对象学的还好吗？？")
    me.save_datas_to_file("datas.xlsx")

