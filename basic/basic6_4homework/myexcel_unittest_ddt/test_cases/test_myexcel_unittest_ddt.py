

# 对excel类进行单元 测试
# 引入ddt，准备多组数据。

# 1、我要设计什么样的测试用例？
  # 用例1  读取第3,4,5行数据。断言：数据不空。
  # 用例2  读取单元格(2,3)(3,4)的数据。断言：(2,3)与长沙是否相等。
  # 用例3  写入单元格(6,8)(8,8)的值：666,520  断言：是否写入成功。


import unittest
from openpyxl import load_workbook
from basic.basic6_4homework.myexcel_unittest_ddt.tools.my_excel import MyExcel
import ddt
import os

@ddt.ddt
class TestExcel(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        # 准备工作
        # homework_0518文件夹的路径
        base_dir = os.path.split(os.path.split(os.path.abspath(__file__))[0])[0]
        # 实例化excel类，得到要操作的表单对象
        cls.me = MyExcel(os.path.join(base_dir,"test_dates/datas.xlsx"))
        cls.me.select_sheet_by_name("case_datas")

    # @classmethod
    # def tearDownClass(cls) -> None:
    #     # 保存写入的所有数据
    #     cls.me.save("datas.xlsx")

    @ddt.data(2,3,4)
    def test_read_row_datas(self,row):
        # 测试数据是什么？2,3,4
        # 步骤是什么?
        datas = self.me.get_dada_by_row(row)
        # 断言是什么？
        self.assertIsNotNone(datas)

    # 测试数据
    cell_datas = [(2,3,"长沙"),(3,4,"分享"),(1,1,"姓名")]

    @ddt.data(*cell_datas)
    def test_read_cell_data(self,data):
        # 测试数据是什么？2,3
        # 步骤是什么?
        datas = self.me.get_data_by_cell(data[0],data[1])
        # 断言是什么
        self.assertEqual(data[2],datas)

    # 测试数据
    new_datas = [(6,8,"666"),(8,8,"520")]

    @ddt.data(*new_datas)
    @ddt.unpack
    def test_write_cell_data(self,row,col,new_value):
        # 测试数据是什么？6,8,666
        # 步骤是什么?
        self.me.write_data_by_cell(row,col,new_value)
        self.me.save_datas("datas.xlsx")
        # 断言是什么
        # 1、从excel当中读取单元格6，8的值。然后再与666比对是否相等。
        wb = load_workbook("datas.xlsx")
        self.assertEqual(new_value,wb["case_datas"].cell(row,col).value)





if __name__ == "__main__":
    unittest.main()