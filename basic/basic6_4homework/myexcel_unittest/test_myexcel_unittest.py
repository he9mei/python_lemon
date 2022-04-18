#!/usr/bin/python3
# -*- coding: utf-8 -*-
# Name: class_excel_unittest
# Author: 简
# Time: 2019/5/16

# 对excel类进行单元 测试

# 1、我要设计什么样的测试用例？
  # 用例1  读取第3行数据。断言：数据不空。
  # 用例2  读取单元格(2,3)的数据。断言：(2,3)与长沙是否相等。
  # 用例3  写入单元格(6,8)的值：666  断言：是否写入成功。

import unittest
from openpyxl import load_workbook
from basic.basic6_4homework.myexcel_unittest.my_excel import MyExcel


class TestExcel(unittest.TestCase):   # 测试类没有init方法

    @classmethod
    def setUpClass(cls):
        # 准备工作
        # 实例化excel类，得到要操作的表单对象
        cls.me = MyExcel("datas.xlsx")
        cls.me.select_sheet_by_name("case_datas")

    # @classmethod
    # def tearDownClass(cls) -> None:
    #     # 保存写入的所有数据
    #     cls.me.save("datas.xlsx")

    def test_read_row_datas(self):
        # 测试数据是什么？3
        # 步骤是什么?
        datas = self.me.get_dada_by_row(3)
        # 断言是什么？
        self.assertIsNotNone(datas)


    def test_read_cell_data(self):
        # 测试数据是什么？2,3
        # 步骤是什么?
        datas = self.me.get_data_by_cell(2,3)
        # 断言是什么
        self.assertEqual("长沙",datas)

    def test_write_cell_data(self):
        # 测试数据是什么？6,8,666
        # 步骤是什么?
        self.me.write_data_by_cell(6,8,"666")
        self.me.save_datas("datas.xlsx")
        # 断言是什么
        # 1、从excel当中读取单元格6，8的值。然后再与666比对是否相等。
        wb = load_workbook("datas.xlsx")
        self.assertEqual("666",wb["case_datas"].cell(6,8).value)

    # 引入ddt，准备多组数据。


if __name__ == "__main__":
    unittest.main()