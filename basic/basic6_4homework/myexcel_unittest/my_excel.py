#!/usr/bin/python3
# -*- coding: utf-8 -*-
# Name: my_excel
# Author: liyuan
# Time: 16:11


"""
第一：excel类封装需要提供以下功能：
1、选择表单功能
2、读取一个单元格的数据功能
3、读取一行数据  功能
4、读取表单中所有数据功能
5、往单元格中写入数据功能
6、保存数据功能

第二：请设计测试数据，对封装的excel类功能进行测试。
excel中的具体内容，由各位同学自由发挥
"""
from openpyxl import load_workbook

class MyExcel:

    # 加载excel文件。
    def __init__(self,filepath):
        # 文件存在就加载。不存在就报错。
        try:
            self.wb = load_workbook(filepath)
        except:
            print("加载excel文件失败！！请检查！")
            raise


    # 选择表单功能
    def select_sheet_by_name(self,sheet_name="Sheet1"):
        if sheet_name in self.wb.sheetnames:
            self.sh = self.wb[sheet_name]
        else:
            print("表单名称，在当前excel文件中不存在，请检测表单名称！")


    # 读取一个单元格的数据功能
    def get_data_by_cell(self,row,column):
        # 判断行号、列号有效
        if self._check_num_valid(row,self.sh.max_row) is True and \
                self._check_num_valid(column,self.sh.max_column) is True:
            return self.sh.cell(row,column).value

    # 读取一行数据功能
    def get_dada_by_row(self,row):
        row_datas = []
        # 判断行号有效
        if self._check_num_valid(row, self.sh.max_row) is True:
            # 有则读取一行数据
            for col in range(1,self.sh.max_column+1):
                row_datas.append(self.sh.cell(row,col).value)
        return row_datas

    # 读取所有数据
    def get_data_all(self):
        all_datas = {}
        for row in range(1,self.sh.max_row + 1):
            row_datas = self.get_dada_by_row(row)
            all_datas["第{}行: ".format(row)] = row_datas
        return all_datas

    # 写入数据功能
    def write_data_by_cell(self,row,column,value):
        self.sh.cell(row,column).value = value

    # 保存功能
    def save_datas(self,filepath):
        try:
            self.wb.save(filepath)
        except PermissionError:
            print("要操作的文件，没有写入权限。请检查权限！")
            raise
        except FileNotFoundError:
            print("文件路径不存在，请确保路径正确！！")
            raise
        except:
            print("保存写入的数据失败！！请检查异常")
            raise

    # 检测数据有效
    def _check_num_valid(self,cur_num,max_num):
        # 类型检测
        if type(cur_num) is not int and type(cur_num) is not str:
            print("cur_num 数据类型错误！请确认为整数类型，或者为字符串类型！")
            return
        # 数字检测
        if type(cur_num) is str:
            try:
                cur_num = int(cur_num)
            except:
                print("cur_num参数非法！请确认是输入数据为整数数字！")
                return
        # 数字是否出范围
        if cur_num in range(1,max_num+1):
            return True
        else:
            print("行号或者列号，超出了目前最大行号，或者最大列号！！")
            return False


if __name__ == '__main__':

    # 对自己定义的功能类进行测试
    # 正常功能测试
    e = MyExcel("datas.xlsx") # 加载文件
    e.select_sheet_by_name("case_datas")  # 选表单
    print(e.get_dada_by_row(3)) # 读取第三行数据
    print(e.get_data_by_cell(2,3)) # 读取某一个数据
    e.write_data_by_cell(6,6,"我悄咪咪的进来了！") # 写入数据
    e.write_data_by_cell(7,7,"哼哼，我也悄悄进来了！") # 写入数据
    e.save_datas("datas.xlsx")  # 保存数据

    # 异常功能测试
    # 1、excel文件路径不存在
    # e = MyExcel("today/datas.xlsx") # 加载文件
    # 2、表单名称不存在
    # e = MyExcel("datas.xlsx") # 加载文件
    # e.select_sheet_by_name("Sheet1")  # 选表单
    # 3、读取的行号和列号，超出了范围
    # e = MyExcel("datas.xlsx") # 加载文件
    # e.select_sheet_by_name("case_datas")  # 选表单
    # e.get_data_by_cell(10,10)
