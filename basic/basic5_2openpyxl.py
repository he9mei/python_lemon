#!/usr/bin/python3
# -*- coding: utf-8 -*-

'''
openpyxl---针对excel后缀名为xlsx操作的第三方库
'''

# ---老师的笔记---

# 第三方库 openpyxl

# install 安装
# pip install openpyxl

# 引入第三方库
# excel操作的流程：
# 打开excel，进入工作薄 workbook
# 选择表单 Sheet
# 单元格  Cell
# 读写操作

# 测试数据是已经存在的。表格至少是存在的。

from openpyxl import load_workbook
wb = load_workbook(r'D:\Pychram-Workspace\python17\class_20190507\datas.xlsx')
# wb.read_only
# from openpyxl.workbook import Workbook

# 表单 workSheet
sh = wb["case_datas"]

# 获取单元格-值. 从1开始。# 读取单元格的值
print(sh.cell(1,1).value)

# # 写单元格的值
# sh.cell(6,1).value = "二狗"

# # 保存写入的数据 - 整个工作薄
# wb.save(r'D:\Pychram-Workspace\python17\class_20190507\datas.xlsx')


# 总行号
rows = sh.max_row
print(rows)
colums = sh.max_column
print(colums)

# 遍历行号
for row in range(2,rows+1):  # 行号
    print("第几行：",row)
    row_datas = {}
    for col in range(1,colums+1):  # 列号
        # print(sh.cell(row,col).value)
        # 第一行不动。第一行的所有列都是key  key对应的value是谁？
        row_datas[sh.cell(1,col).value] = sh.cell(row,col).value
    print("本行的数据为：",row_datas)


# 姓名：小简  年龄：20   班级：py17
# 第一行不动。第一行的所有列都是key  key对应的value是谁？

# 封装为一个类。exel样式内容。读(一行、所有的数据)、写操作、保存操作
# 测试自己的类，是否功能有bug。

# pip install pandas

# today 总结
# 反射 hasattr  getattr  setattr delattr
# exel数据操作  - openpyxl
# workbook、sheet、cell
# 加载一个工作薄  wb = load_workbook(excel_filepath)  r
# 挑表单  sh = wb["表单名称"]
# 挑单元格   sh.cell(row,column) # 对象
# 读取单元格  value = sh.cell(row,column).value  # 下标从1开始
# 修改/添加值  sh.cell(row,column).value = new_value
# 保存   wb.save(excel_filepath) # 由路径 决定 是另存为，还是原来的文件。
# 总行号  sh.max_row
# 总列号  sh.max_column

# 读取所有的数据  for for









