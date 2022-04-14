#!/usr/bin/python3
# -*- coding: utf-8 -*-

'''
读取单元格练习
'''
# import os
# print(os.getcwd())
import openpyxl
wb = openpyxl.load_workbook(r'C:/Users/lipan/PycharmProjects/python_lemon/basic/testData.xlsx')
# 遇到问题，
# 直接在pycharm新建的文件data.xlsx报错：ipfile.BadZipFile: File is not a zip file
# 从wps新建的文件可以
sh = wb['data']
cell_value = sh.cell(row=1,column=1).value
# print(cell_value)

list_case=[]
for i in range(2,sh.max_row+1):  # 第1行作为表头，从第2行开始遍历
    case={}
    for j in range(1,sh.max_column+1):
        keys = sh.cell(1, j).value
        case[keys]=sh.cell(i,j).value
    # print(case)
    list_case.append(case)
print(list_case)
'''
{'姓名': '何丫丫', '年龄': 20, '城市': '上海'}
{'姓名': '李晓霞', '年龄': 28, '城市': '北京'}
{'姓名': '王二狗', '年龄': 22, '城市': '重庆'}

[{'姓名': '何丫丫', '年龄': 20, '城市': '上海'},
 {'姓名': '李晓霞', '年龄': 28, '城市': '北京'}, 
 {'姓名': '王二狗', '年龄': 22, '城市': '重庆'}]
'''


# 作业：
# 类中方法:读、写并保存操作
class DoExcel:
    def __init__(self,file_path,sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.workbook = openpyxl.load_workbook(file_path)
        self.sheet = self.workbook[self.sheet_name]

    def read_excel_cell(self,row,column):
        cell_value = self.sheet.cell(row,column).value
        self.workbook.close()
        return cell_value

    def read_all_list_dict(self):
        list_case = []
        for i in range(2,self.sheet.max_row+1):
            case={}
            for j in range(1,self.sheet.max_column+1):
                key=self.sheet.cell(1,j).value
                value=self.sheet.cell(i,j).value
                case[key]=value
            list_case.append(case)
        self.workbook.close()
        return list_case

    def wirte(self,row,column,data):
        self.sheet.cell(row,column).value=data
        self.workbook.save(self.file_path)
        self.workbook.close()


file_path=r'C:/Users/lipan/PycharmProjects/python_lemon/basic/testData.xlsx'
excel = DoExcel(file_path,'data')
excel.wirte(4,1,'小白')
cases = excel.read_all_list_dict()
print(cases)